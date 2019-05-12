#!/usr/bin/env python
# -*- coding: UTF-8 -*-
from datetime import datetime
import os
import dailyreport_data
import dailyreport_escode
from multiprocessing import Pool
from openpyxl import load_workbook
from collections import ChainMap


def callback(result):
    global excel_value
    excel_value = ChainMap(excel_value,result)
    return


def excel_analyze(landscape):
    print(landscape[0], 'start', 'sub PID:', os.getpid())
    data = dailyreport_data.Data(landscape[0])
    escode1 = dailyreport_escode.Analyze(landscape[0])
    escode2 = dailyreport_escode.EsCode(landscape[0])
    access_host = data.analyze(escode1.es_access_host, escode2.index1, landscape[1], landscape[2])
    if landscape[0] in ('CN', 'EU'):
        occ_tenant = data.analyze(escode1.es_occ_tenant, escode2.index2, landscape[1], landscape[4])
        access_http_agent = data.analyze(escode1.es_access_http_agent, escode2.index1, landscape[1], landscape[3])
        job_tenant = data.analyze(escode1.es_job_tenant, escode2.index2, landscape[1], landscape[5])
    elif landscape[0] == 'MSA':
        access_http_agent = data.analyze(escode1.es_access_http_agent, escode2.index1, landscape[1], landscape[3])
        common = data.analyze(escode1.es_common, escode2.index2, landscape[1], landscape[4])
        common_level = data.analyze(escode1.es_common_level, escode2.index2, landscape[1], landscape[5])
        service = data.service_analyze(escode1.es_service, escode2.index2, landscape[1], landscape[6])
    else:
        pass
    if landscape[0] in ('CN', 'EU'):
        data_dict = ChainMap(access_host, access_http_agent, occ_tenant, job_tenant)
    else:
        data_dict = ChainMap(access_host, access_http_agent, common, common_level, service)
    print(landscape[0], 'Done')
    return data_dict


excel_value = {}

cn = ['CN', 3, ('a', 'b', 'c'), ('f', 'g', 'h'), ('k', 'l', 'm'), ('p', 'q', 'r')]
# us = ['US', 16, ('a', 'b', 'c'), ('f', 'g', 'h'), ('k', 'l', 'm'), ('p', 'q', 'r')]
# eu = ['EU', 29, ('a', 'b', 'c'), ('f', 'g', 'h'), ('k', 'l', 'm'), ('p', 'q', 'r')]
msa = ['MSA', 42, ('a', 'b', 'c'), ('f', 'g', 'h'), ('k', 'l', 'm'), ('p', 'q', 'r'), ('u', 'v', 'w', 'x', 'y')]
landscapes = (cn,  msa)
# landscapes = (msa,)

if __name__ == "__main__":
    print(datetime.now(), 'master PID:', os.getpid())
    pool = Pool()
    for land in landscapes:
        pool.apply_async(excel_analyze, (land,), callback=callback)
        # pool.apply_async(excel_analyze(land))
    pool.close()
    pool.join()
    print(excel_value)
    wb = load_workbook('analyze_template.xlsx')
    ws = wb.active
    for key in excel_value:
        ws[key] = excel_value[key]
    wb.save('analyze.xlsx')
    print(datetime.now(), 'Analyze is done.')
