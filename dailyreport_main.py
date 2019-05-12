#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import dailyreport_data
import multiprocessing
from multiprocessing import Pool
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import configparser
import os
from collections import ChainMap

cf = configparser.ConfigParser()
cf.read('config.ini')

cn = ['CN', ('a', 'c'), ('b', 'c'), ('b', 'c', 'd'), (('a', 'b'), ('c', 'd'))]
# eu = ['EU', ('e', 'g'), ('f', 'g'), ('f', 'g', 'h'), (('e', 'f'), ('g', 'h'))]
# msa = ['MSA', ('i', 'k'), ('j', 'k'), ('j', 'k', 'l'), (('i', 'j'), ('k', 'l'))]
msa = ['MSA', ('e', 'g'), ('f', 'g'), ('f', 'g', 'h'), (('e', 'f'), ('g', 'h'))]

excel_value = {}


def excel_http(http_data, a):
    dict_http = {}
    http_status = http_data["aggregations"]['2']['buckets']
    for x in range(0, len(http_status)):
        code = http_status[x]['key']
        count = http_status[x]['doc_count']
        l1 = a[0]
        l2 = a[1]
        http_row = 12 + x
        http_row = str(http_row)
        coderow = l1 + http_row
        countrow = l2 + http_row
        dict_http[coderow] = code
        dict_http[countrow] = count
    return dict_http


def excel_hana(a, hana_status):
    dict_hana = {}
    for q in (0, 1, 2):
        line = a[q]
        coderow = line + str(45)
        dict_hana[coderow] = hana_status[q]
    return dict_hana


def excel_pv(a, count):
    dict_pv = {}
    for k in (0, 1):
        pv_row = a[k] + str(35)
        dict_pv[pv_row] = count[k]
    return dict_pv


def excel_log(a, count):
    dict_log = {}
    rows = []
    for l in (0, 1):
        for m in range(0, int(len(count) / 2)):
            log_row = a[l] + str(39 + m)
            rows.append(log_row)
    for n in range(0, len(rows)):
        dict_log[rows[n]] = count[n]
    return dict_log


def excel_es_status(a, count):
    dict_es_status = {}
    for m in range(0, len(count)):
        if count[m] == '0':
            status = 'red'
        else:
            status = 'green'
        rowname = a[0] + str(63 + m)
        dict_es_status[rowname] = status
    return dict_es_status


def excel_s3(status):
    dict_s3 = {}
    scape = [cn[2], msa[2]]
    for o in range(0, len(scape)):
        backup_status = status[o]
        for q in range(0, len(backup_status)):
            rowb = scape[o][0] + str(47 + q)
            rowc = scape[o][1] + str(47 + q)
            dict_s3[rowb] = backup_status[q]
            dict_s3[rowc] = backup_status[q]
    return dict_s3


def dailyreport(landscape):
    print(landscape[0], 'sub PID:', os.getpid())
    data = dailyreport_data.Data(landscape[0])
    print("Get " + landscape[0] + " hana license and local backup status")
    hana_status = data.hana()
    dict_hana = excel_hana(landscape[3], hana_status)
    print("Get %s hana disk status" % landscape[0])
    # dict_disk_status = data.hana_disk_check(landscape[4])
    if landscape[0] == 'MSA':
        print("Get s3 backup status!")
        s3_status = data.s3_backup()
        dict_backup = excel_s3(s3_status)
    print("Get " + landscape[0] + " http status")
    http_data = data.http()
    dict_http = excel_http(http_data, landscape[1])
    print("Get " + landscape[0] + " pv")
    pv_count = data.pv()
    dict_pv = excel_pv(landscape[1], pv_count)
    print("Get " + landscape[0] + " log")
    if landscape[0] == 'MSA':
        log_count = data.log(1)
        es_count = data.check_es_index(1)
    else:
        log_count = data.log(0)
        es_count = data.check_es_index(0)
    dict_log = excel_log(landscape[2], log_count)
    print("Get %s es status" % landscape[0])
    dict_es_status = excel_es_status(landscape[2], es_count)
    if landscape[0] == 'MSA':
        # dict_all = ChainMap(dict_http, dict_pv, dict_hana, dict_log, dict_backup, dict_es_status, dict_disk_status)
        dict_all = ChainMap(dict_http, dict_pv, dict_hana, dict_log, dict_backup, dict_es_status)
    else:
        # dict_all = ChainMap(dict_http, dict_pv, dict_hana, dict_log, dict_es_status, dict_disk_status)
        dict_all = ChainMap(dict_http, dict_pv, dict_hana, dict_log, dict_es_status)
    print(landscape[0] + ' done')
    return dict_all


def callback(result):
    global excel_value
    excel_value = ChainMap(excel_value, result)
    return


def shift_info():
    hour = datetime.utcnow().hour
    if hour <= 7:
        shift = 'Morning'
    elif hour <= 15:
        shift = 'Middle'
    else:
        shift = 'Night'
    excel_value['e3'] = shift
    user = str(cf.get('people', 'shift_peope'))
    excel_value['e4'] = user


landscapes = (cn, msa)
# landscapes = (msa,)
wb = load_workbook('dailyreport_template.xlsx')
ws = wb.active

if __name__ == "__main__":
    print(datetime.now())
    shift_info()
    pool = Pool()
    multiprocessing.freeze_support()
    for land in landscapes:
        scope = land
        pool.apply_async(dailyreport, (scope,), callback=callback)
        # pool.apply_async(dailyreport(scope), callback=callback)
    pool.close()
    pool.join()

    for value_key in excel_value.keys():
        ws[value_key] = excel_value[value_key]
    bd = Side(style='thin', color='000000')
    for j in range(97, 105):
        for i in range(1, 62):
            row = chr(j) + str(i)
            p = ws[row]
            p.border = Border(top=bd, left=bd, right=bd, bottom=bd)
    wb.save('dailyreport.xlsx')
    print(datetime.now())
    print("Finished")
