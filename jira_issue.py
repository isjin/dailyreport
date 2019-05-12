#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import requests
import json
import time
from datetime import datetime
import os
from openpyxl import load_workbook
from jira import JIRA
import base64



def jira():
    options = {
        'server': 'https://ubtjira.smec.sap.corp',
        'verify': False,
    }
    jira_api = JIRA(basic_auth=('c5229570', 'Initial2'), options=options)
    return jira_api


def shift_info():
    hour = datetime.utcnow().hour
    if hour <= 7:
        shift = 'Morning Shift'
    elif hour <= 15:
        shift = 'Middle Shift'
    else:
        shift = 'Night Shift'
    return shift


def create_dailyreport_jira(summary, comment):
    data = {"fields": {"project": {"key": "DEVOPS"}, "summary": summary, "description": comment,
                       "assignee": {"name": 'c5229570'}, "issuetype": {"id": "11600"}}}
    data_json = json.dumps(data)
    return data_json


def create_issue_jira(summary, comment):
    data = {"fields": {"project": {"key": "DEVOPS"}, "summary": summary, "description": comment,
                       "assignee": {"name": 'c5229570'}, "issuetype": {"id": "10100"}}}
    data_json = json.dumps(data)
    return data_json


def request(jira_api_url, data):
    response = requests.post(jira_api_url, headers=headers, verify=False, data=data)
    issue_data = response.text
    print(issue_data)
    data = json.loads(issue_data)
    return data


jira_issue_api = 'https://ubtjira.smec.sap.corp/rest/api/2/issue'
jira_issue_link_api = 'https://ubtjira.smec.sap.corp/rest/api/2/issueLink'
headers = {
    'Content-Type': 'application/json',
    'cache-control': "no-cache",
    'authorization': 'Basic YzUyMjk1NzA6SW5pdGlhbDE=',
}
if __name__ == "__main__":
    jira = jira()
    m_time = os.path.getmtime('analyze.xlsx')
    date = time.localtime(m_time)
    file_m_date = time.strftime('%Y%m%d', date)
    today = time.strftime('%Y%m%d', time.localtime(time.time()))
    title = '[%s][%s]' % (today, shift_info())
    daily_report = 'dailyreport.xlsx'
    email = r'C:/Users/c5229570/Desktop/Daily Report %s %s.msg' % (today, shift_info())
    issue_summary = jira.search_issues('issuetype=11600 and reporter=c5229570', maxResults=1)[0].fields.summary
    if today not in issue_summary:
        if today != file_m_date:
            description = "Everything is fine!"
            create_json = create_dailyreport_jira(title, description)
            jira_data = request(jira_issue_api, create_json)
            # jira_data = {"id": "249817", "key": "DEVOPS-4369","self": "https://ubtjira.smec.sap.corp/rest/api/2/issue/249817"}
            issue = jira_data['key']
            jira.add_attachment(issue, attachment=daily_report)
            # jira.add_attachment(issue, attachment=email)
            #jira.add_comment(issue,"CN hana backup to local and s3 failed.")
            jira.transition_issue(issue, '2')
        else:
            wb1 = load_workbook('dailyreport.xlsx')
            ws1 = wb1['test']
            cn = ['CN', ('a', 'b', 'c'), ]
            # eu = ['EU', ('e', 'f', 'g'), ]
            # eu = ['EU', ('i', 'j', 'k'), ]
            msa = ['MSA', ('e', 'f', 'g'), ('k', 'l', 'm')]
            landscape = (cn, msa)
            notes = []
            out_issues = []
            for i in landscape:
                if i[0] != 'MSA':
                    number = 3
                else:
                    number = 2
                for j in range(0, number):
                    row_service = str(i[1][0]) + str(j + 39)
                    row_24 = str(i[1][1]) + str(j + 39)
                    row_48 = str(i[1][2]) + str(j + 39)
                    service = str(ws1[row_service].value)
                    count_24 = float(ws1[row_24].value)
                    count_48 = float(ws1[row_48].value)
                    rate = (count_24 - (count_48 - count_24)) / (count_48 - count_24) * 100
                    rate = round(rate, 2)
                    if rate > 50:
                        note = i[0] + ' ' + service + " growth rate is high." + '\n'
                        notes.append(note)
                if i[0] == 'MSA':
                    wb2 = load_workbook('analyze.xlsx')
                    ws2 = wb2['Sheet1']
                    for l in range(0, 10):
                        service_row = str(i[2][0]) + str(l + 42)
                        h24_row = str(i[2][1]) + str(l + 42)
                        h48_row = str(i[2][2]) + str(l + 42)
                        service_name = str(ws2[service_row].value)
                        try:
                            count_h24 = float(ws2[h24_row].value)
                            count_h48 = float(ws2[h48_row].value)
                        except Exception as e:
                            print(e.__str__())
                            continue
                        if count_h48 is not None:
                            service_rate = (count_h24 - (count_h48 - count_h24)) / (count_h48 - count_h24) * 100
                            service_rate = round(service_rate, 2)
                            if service_rate > 250:
                                jira_title = service_name + ' logs increased rapidly'
                                jira_description = '''{code}serivice   24hours    48hours    rate\n%s         %s         %s         %s%%{code}
                                        ''' % (service_name, count_h24, count_h48, service_rate)
                                service_jira_json = create_issue_jira(jira_title, jira_description)
                                service_jira_data = request(jira_issue_api, service_jira_json)
                                service_issue = service_jira_data['key']
                                # jira.assign_issue(service_issue, 'c5229570')
                                out_issues.append(service_issue)
            description = ''
            for k in range(0, len(notes)):
                description = description + notes[k]
            dailyreport_json = create_dailyreport_jira(title, description)
            jira_data = request(jira_issue_api, dailyreport_json)
            issue = jira_data['key']
            jira.add_attachment(issue, attachment=daily_report)
            # jira.add_attachment(issue, attachment=email)
            for out_issue in out_issues:
                jira.create_issue_link('Relates', issue, out_issue)
            jira.add_comment(issue, 'Add all sub links.')
            # jira.transition_issue(issue,'2')
            print (jira_data)
